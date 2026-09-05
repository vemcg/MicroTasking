#!/usr/bin/env python3
# Copyright (c) 2026 Vern McGeorge. All rights reserved.
"""Builds the Google Sheets import template (one tab per category) from content/tasks.json.

Upload the resulting .xlsx to Google Drive; it opens as a native Sheet with tabs already
matching the external-task-source schema (description, link). Duration is app-maintained
and adaptive, so it's never authored in the sheet.
"""
import argparse
import json
import pathlib

from openpyxl import Workbook
from openpyxl.worksheet.hyperlink import Hyperlink

#!/usr/bin/env python3
"""Builds the Google Sheets import template (one README tab + category tabs) from content/tasks.json.

Upload the resulting .xlsx to Google Drive; it opens as a native Sheet with tabs already
matching the external-task-source schema (checkbox in col A, description in col B, link in col C).
"""
import argparse
import json
import pathlib

from openpyxl import Workbook
from openpyxl.worksheet.hyperlink import Hyperlink

INVALID_SHEET_TITLE_CHARS = str.maketrans({c: "-" for c in "/\\?*[]:"})

README_CONTENT = [
    ["MICROTASKING TASK POOL TEMPLATE"],
    [""],
    ["Welcome to your MicroTasking Task Pool spreadsheet!"],
    [""],
    ["HOW TO USE THIS SPREADSHEET:"],
    ["1. CATEGORIES (TABS): Each tab at the bottom represents a category (e.g. Decluttering, Cleaning, Paperwork, Finances, Health, Errands)."],
    ["   - You can add new tabs, rename existing tabs, or delete tabs you don't need."],
    [""],
    ["2. COLUMNS IN TASK TABS:"],
    ["   - Column A (Master Toggle / Enabled): Cell A1 toggles the whole tab, and each task row below also has its own independent checkbox."],
    ["   - Column B (Description): The text description of the micro-task."],
    ["   - Column C (Link): Optional URL (e.g. video tutorial, document, or web tool)."],
    [""],
    ["3. SYNCING WITH THE APP:"],
    ["   - Set Share permissions to 'Anyone with the link can view'."],
    ["   - Paste your Sheet URL into the onboarding page to generate your custom QR code."],
    ["   - In the MicroTasking app, tap Settings -> Import External Task Pool -> Scan QR Code."]
]


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--tasks-json", default="content/tasks.json")
    parser.add_argument("--out", default="content/microtasking-sheet-template.xlsx")
    args = parser.parse_args()

    data = json.loads(pathlib.Path(args.tasks_json).read_text(encoding="utf-8"))

    workbook = Workbook()
    workbook.remove(workbook.active)

    # 1. First tab: README (in ALL CAPS)
    readme_sheet = workbook.create_sheet(title="README")
    for row in README_CONTENT:
        readme_sheet.append(row)
    readme_sheet.column_dimensions["A"].width = 110

    # 2. Category tabs
    for category in data["categories"]:
        raw_name = category["name"]
        # Exception requested by user: rename "Admin / Paperwork" to "Paperwork"
        if "Admin" in raw_name or "Paperwork" in raw_name:
            raw_name = "Paperwork"

        title = raw_name.translate(INVALID_SHEET_TITLE_CHARS)[:31]
        sheet = workbook.create_sheet(title=title)
        
        # Row 1: A1 = master toggle, B1 = description, C1 = link
        sheet.cell(row=1, column=1, value=True)
        sheet.cell(row=1, column=2, value="description")
        sheet.cell(row=1, column=3, value="link")

        for row_index, task in enumerate(category["tasks"], start=2):
            # Each task row has its own independent checkbox value, while A1 remains the master toggle.
            sheet.cell(row=row_index, column=1, value=True)
            sheet.cell(row=row_index, column=2, value=task["description"])
            link = task.get("link", "")
            link_cell = sheet.cell(row=row_index, column=3, value=link or None)
            if link:
                link_cell.hyperlink = Hyperlink(ref=link_cell.coordinate, target=link)

        sheet.column_dimensions["A"].width = 12
        sheet.column_dimensions["B"].width = 75
        sheet.column_dimensions["C"].width = 35

    pathlib.Path(args.out).parent.mkdir(parents=True, exist_ok=True)
    workbook.save(args.out)
    print(f"Wrote {args.out}")


if __name__ == "__main__":
    main()



if __name__ == "__main__":
    main()
